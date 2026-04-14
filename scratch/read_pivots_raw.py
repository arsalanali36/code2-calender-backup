import pandas as pd
import os

file_path = r'D:\KHAZANA\KHAZANA\PYTHON\CODE2- CALENDER\PINE SCRIPTS\PIVOT LEVELS.xlsx'
if os.path.exists(file_path):
    try:
        # Try reading with different engines or just openpyxl directly
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"--- Sheet: {sheet} ---")
            for row in ws.iter_rows(values_only=True):
                print(row)
    except Exception as e:
        print(f"Error: {e}")
else:
    print("File not found.")
