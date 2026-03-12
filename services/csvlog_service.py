"""
services/csvlog_service.py
--------------------------
Parse LOGGER.xlsx schema and return structured JSON for the CSVLog modal.
No Flask imports, no request/response objects.
"""
import os
import re


def _to_key(head):
    return re.sub(r'^_|_$', '', re.sub(r'[^a-z0-9]+', '_', head.lower()))


def load_schema(schema_file):
    """
    Parse the csvlog_schema.xlsx file.
    Returns dict:
        {
          "groups": ["Zone", "Entry", "Exit", "PSy"],
          "fields": {
            "Zone": [
              { "head": "formed", "type": "Switch", "input": "Y/N", "options": ["Y", "N"] },
              ...
            ],
            ...
          }
        }
    Only rows with Display == "Show" are included.
    Returns None if file doesn't exist.
    Returns {"error": "..."} on parse failure.
    """
    if not os.path.exists(schema_file):
        return None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(schema_file, data_only=True)
        ws = wb.active

        groups_order = []
        fields = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Columns: Group, Head, Input, Type, Display, Description
            padded = list(row) + [None] * 6
            group, head, input_val, type_val, display = (
                padded[0], padded[1], padded[2], padded[3], padded[4]
            )

            if not group or not head:
                continue
            if str(display or '').strip().lower() != 'show':
                continue

            group    = str(group).strip()
            head     = str(head).strip()
            type_val = str(type_val or '').strip()
            input_val = str(input_val).strip() if input_val is not None else None

            if group not in fields:
                groups_order.append(group)
                fields[group] = []

            # Parse options list for Dropdown / Range types
            options = None
            if input_val and type_val in ('Dropdown', 'Range'):
                # Strip surrounding quotes, normalise newlines → commas
                clean = input_val.strip('"').replace('\n', ',')
                options = [o.strip() for o in clean.split(',') if o.strip()]

            # Y/N Switch → options list
            if type_val == 'Switch' and input_val == 'Y/N':
                options = ['Y', 'N']

            fields[group].append({
                'head':    head,
                'type':    type_val,
                'input':   input_val,
                'options': options,
            })

        return {'groups': groups_order, 'fields': fields}

    except Exception as exc:
        return {'error': str(exc)}


def export_csvlog_excel(trades, schema_file=None):
    """
    Export all trades with csvlog data to an Excel BytesIO.
    Returns (BytesIO, None) on success, (None, error_str) on failure.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
        from io import BytesIO
    except ImportError:
        return None, 'openpyxl not installed'

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'CSVLog Export'

        # ── Build ordered list of csvlog columns ──────────────────────────────
        # (group_key, field_key, display_label)
        csvlog_cols = []
        seen_cols = set()

        schema = load_schema(schema_file) if schema_file and os.path.exists(schema_file) else None
        if schema:
            for group in schema.get('groups', []):
                gk = group.lower()
                section = ''
                for field in schema['fields'].get(group, []):
                    if field['type'] == '-':
                        section = _to_key(field['head'])
                        continue
                    fk = (_to_key(section + '_' + field['head']) if section
                          else _to_key(field['head']))
                    label = f"{group} / {field['head']}"
                    pair = (gk, fk)
                    if pair not in seen_cols:
                        seen_cols.add(pair)
                        csvlog_cols.append((gk, fk, label))

        # Explicitly add Body Vitals columns with nice labels
        _vitals_keys = [
            ('alertness', 'Body Vitals / Alertness'),
            ('neend',     'Body Vitals / Neend (Sleep)'),
            ('potty',     'Body Vitals / Potty'),
            ('sabar',     'Body Vitals / Sabar vs Impulsive'),
        ]
        for fk, label in _vitals_keys:
            pair = ('body_vitals', fk)
            if pair not in seen_cols:
                seen_cols.add(pair)
                csvlog_cols.append(('body_vitals', fk, label))

        # Pick up any extra keys from actual data
        for trade in trades:
            for gk, fdata in (trade.get('csvlog') or {}).items():
                if not isinstance(fdata, dict):
                    continue
                for fk in fdata:
                    if fk.endswith('_obs') or fk == '_meta':
                        continue
                    pair = (gk, fk)
                    if pair not in seen_cols:
                        seen_cols.add(pair)
                        csvlog_cols.append((gk, fk, f"{gk} / {fk}"))

        # ── Headers ───────────────────────────────────────────────────────────
        base_hdrs = ['Date', 'Instrument', 'TradeType', 'Qty',
                     'Entry Time', 'Exit Time', 'P/L (Rs)', 'Points', 'Note']
        obs_hdr = ['Observations']
        all_hdrs = base_hdrs + [c[2] for c in csvlog_cols] + obs_hdr
        ws.append(all_hdrs)

        # Style header row
        hdr_fill = PatternFill('solid', fgColor='1E2535')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        # ── Data rows ─────────────────────────────────────────────────────────
        def _pick(t, *keys):
            for k in keys:
                v = t.get(k)
                if v is not None and v != '':
                    return v
            return ''

        for trade in trades:
            date      = _pick(trade, 'trade_date', 'Date', 'date')
            instr     = _pick(trade, 'Instrument', 'instrument', 'INSTRUMENT')
            ttype     = _pick(trade, 'TradeType', 'tradetype', 'TRADETYPE')
            qty       = _pick(trade, 'Qty', 'qty', 'QTY')
            buy_t     = _pick(trade, 'Buy Time', 'buy_time', 'BUY TIME')
            sell_t    = _pick(trade, 'Sell Time', 'sell_time', 'SELL TIME')
            pnl       = _pick(trade, 'Net P/L', 'Gross P/L', 'Rs', 'rs', 'RS')
            points    = _pick(trade, 'Pt', 'pt')
            note      = _pick(trade, 'Note', 'note')

            # Entry = earlier time
            try:
                def _tsec(s):
                    p = str(s).split(':')
                    return int(p[0])*3600 + int(p[1])*60 + int(p[2] if len(p)>2 else 0)
                entry_t, exit_t = (buy_t, sell_t) if (_tsec(buy_t) <= _tsec(sell_t)) else (sell_t, buy_t)
            except Exception:
                entry_t, exit_t = buy_t, sell_t

            base_row = [date, instr, ttype, qty, entry_t, exit_t, pnl, points, note]

            csvlog = trade.get('csvlog') or {}
            csvlog_vals = []
            obs_parts = []
            for gk, fk, _ in csvlog_cols:
                gdata = csvlog.get(gk) or {}
                val = gdata.get(fk, '')
                csvlog_vals.append(val)
                obs_val = gdata.get(fk + '_obs', '')
                if obs_val:
                    obs_parts.append(f"[{gk}] {fk.replace('_',' ')}: {obs_val}")

            # Use manually edited obs text if exists
            meta = (csvlog.get('_meta') or {})
            compiled_obs = meta.get('obs_text', '') or '\n'.join(obs_parts)

            ws.append(base_row + csvlog_vals + [compiled_obs])

        # Auto-width columns
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 45)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out, None

    except Exception as exc:
        return None, str(exc)


def generate_logger_template(schema_file=None):
    """
    Generate a protected LOGGER.xlsx template BytesIO.
    - Preserves existing schema rows from schema_file (if available)
    - Appends Body Vitals group rows if not already present
    - Row 1 header is locked; sheet protected to prevent col delete/rename
    - No password → user can unprotect anytime (protects against accidents)
    Returns (BytesIO, None) or (None, error_str).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Protection
        from openpyxl.worksheet.protection import SheetProtection
        from io import BytesIO
    except ImportError:
        return None, 'openpyxl not installed'

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'LOGGER Schema'

        # ── Header row ──────────────────────────────────────────────────────
        headers = ['Group', 'Head', 'Input', 'Type', 'Display', 'Description']
        ws.append(headers)
        hdr_fill = PatternFill('solid', fgColor='1A3050')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            cell.protection = Protection(locked=True)   # lock header cells

        # ── Column widths ────────────────────────────────────────────────────
        for col, w in zip('ABCDEF', [16, 22, 30, 12, 10, 35]):
            ws.column_dimensions[col].width = w

        # ── Load existing schema rows ────────────────────────────────────────
        existing_rows = []
        has_body_vitals = False
        if schema_file and os.path.exists(schema_file):
            try:
                src = openpyxl.load_workbook(schema_file, data_only=True)
                sw = src.active
                for row in sw.iter_rows(min_row=2, values_only=True):
                    vals = list(row) + [None] * 6
                    grp = str(vals[0] or '').strip()
                    if grp.lower() == 'body vitals' or grp.lower() == 'body_vitals':
                        has_body_vitals = True
                    if grp:
                        existing_rows.append(vals[:6])
            except Exception:
                pass

        # Fill existing rows — data cells unlocked so user can edit values
        for r in existing_rows:
            ws.append(r)

        # ── Append Body Vitals if not present ────────────────────────────────
        _VITALS_ROWS = [
            ['Body Vitals', 'Alertness',         '-5,5', 'Range', 'Show', 'Physical alertness level'],
            ['Body Vitals', 'Neend',             '-5,5', 'Range', 'Show', 'Sleep quality last night'],
            ['Body Vitals', 'Potty',             '-5,5', 'Range', 'Show', 'Gut health / comfort'],
            ['Body Vitals', 'Sabar vs Impulsive','-5,5', 'Range', 'Show', 'Patience vs impulsiveness'],
        ]
        if not has_body_vitals:
            ws.append([None] * 6)   # blank spacer row
            for r in _VITALS_ROWS:
                ws.append(r)

        # ── Style data rows: A-D locked (group/head/input/type), E-F editable ─
        # Note: only meaningful when sheet protection is on.
        # A,B = Group,Head locked so names stay stable; C,D,E,F editable.
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                locked = cell.column_letter in ('A', 'B')
                cell.protection = Protection(locked=locked)

        # ── Sheet protection (no password → easy to unprotect intentionally) ──
        ws.protection.sheet          = True
        ws.protection.deleteColumns  = True   # prevent column deletion
        ws.protection.insertColumns  = True   # prevent inserting columns
        ws.protection.sort           = True   # prevent sorting (rearranging)
        # Allow: editing unlocked cells (E=Display, C=Input, D=Type, F=Desc),
        #        inserting rows (adding new fields), selecting any cell.
        ws.protection.insertRows     = False
        ws.protection.deleteRows     = False
        ws.protection.selectLockedCells   = True
        ws.protection.selectUnlockedCells = True

        # ── Instructions sheet ───────────────────────────────────────────────
        ws2 = wb.create_sheet('Instructions')
        ws2.column_dimensions['A'].width = 70
        instructions = [
            ['LOGGER.xlsx — Instructions'],
            [''],
            ['Column guide:'],
            ['  A: Group    — e.g. Zone / Entry / Exit / Body Vitals  (LOCKED — do not rename)'],
            ['  B: Head     — field label shown in the app             (LOCKED — do not rename)'],
            ['  C: Input    — Y/N for Switch; options for Dropdown; min,max for Range'],
            ['  D: Type     — Switch | Input | Dropdown | Range | -    (- = section separator)'],
            ['  E: Display  — Show or Hide'],
            ['  F: Description — your notes (ignored by app)'],
            [''],
            ['Bidirectional sliders: set Input = "-5,5" (or any negative,positive range).'],
            ['  The slider will be centered at 0 and fill green(+) or red(-).'],
            [''],
            ['Conditional freeze rules (built into app, not in schema):'],
            ['  Zone  : If "Zone Created" = N  → Size and Candle fields are frozen'],
            ['  Entry : If "At"  contains "pehle" → Breakout Candle field is frozen'],
            [''],
            ['To add a new field: insert a row below existing ones (allowed).'],
            ['To hide a field:   change Display column from "Show" to "Hide".'],
            ['Columns A-B are locked to prevent accidental renaming.'],
        ]
        for r in instructions:
            ws2.append(r)
        ws2['A1'].font = Font(bold=True, size=13)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out, None

    except Exception as exc:
        return None, str(exc)
